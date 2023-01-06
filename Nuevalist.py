from tkinter import *
import os
import openpyxl
import tkinter as tk
from tkinter import ttk
import sqlite3
from excel1 import excel_lista

def nuevlista():
    raiz=Tk()#creación de ventana
    raiz.title("Nueva Lista")
    raiz.resizable(0,0)
    raiz.iconbitmap("Recursos/Grafico/cheemspng.ico")
    raiz.geometry("1000x650+10+10")
    raiz.config(bg="black")
    raiz.config(bd=1)
    raiz.config(cursor="heart")

    framef=Frame(raiz)
    framef.pack()
    framef.config(bg="#303030")
    framef.config(width="1000",height="650")
    framef.config(bd=1)
    framef.config(cursor="arrow")

    frame1=Frame(raiz)
    frame1.place(x=0, y=0)
    frame1.config(bg="#303030")
    frame1.config(width="1000",height="650")
    frame1.config(bd=1)
    frame1.config(cursor="arrow")

    label1=Label(frame1, text="Ingrese su nueva lista de productos")
    label1.grid(row=0, column=0, sticky="nsew", pady=1, padx=1)
    label1.config(fg="white")
    label1.config(bg="#303030")
    label1.config(font=("Downlink",20))

    frameline=Frame(frame1)
    frameline.grid(row=1, column=0, sticky="nsew", pady=1, padx=20)
    frameline.config(bg="white")
    frameline.config(width="960", height="10")

    frame2=Frame(frame1)
    frame2.grid(row=3, column=0, sticky="nsew", pady=20, padx=20)
    frame2.config(bg="black")
    frame2.config(width="960",height="500")
    frame2.config(bd=1)

    frame3=Frame(frame2)
    frame3.grid(row=0, column=1, sticky="nse", pady=10, padx=100)
    frame3.config(bg="black")
    frame3.config(width="380",height="500")
    frame3.config(bd=1)

    label2=Label(frame3,text="Producto")
    label2.grid(row=0,column=0,pady=2, padx=2)
    label2.config(bg="black", fg='white')
    label2.config(font=("Downlink",12))
    nombre=Entry(frame3)
    nombre.grid(row=0,column=1, pady=2, padx=2)
    nombre.focus()
    
    label3=Label(frame3,text="Cantidad")
    label3.grid(row=1,column=0,pady=2, padx=2)
    label3.config(bg="black", fg='white')
    label3.config(font=("Downlink",12))
    cant=Entry(frame3)
    cant.grid(row=1,column=1,pady=2, padx=2)

    #Mensaje
    mensaje=Label(frame2, text='',bg="black",fg='green',font=("Downlink",8))
    mensaje.grid(row=1,column=1,columnspan=2,sticky=W+E, pady=10, padx=30)


    def queryAlumnos(query):
        #global cur
        conn=sqlite3.connect("Lista.db")
        cur=conn.cursor()
        cur.execute(query)
        conn.commit()
        return cur

    def mostrarDatos():
        registros=treeview.get_children()
        for registro in registros:
            treeview.delete(registro)
        cur=queryAlumnos("SELECT `nombre`,`cant` FROM `Productos`")
        for (nombre,cant) in cur:
            treeview.insert('',0,text=nombre,values=cant)

    def agregarRegistro():
        if len(nombre.get())!=0 and len(cant.get())!=0:
            query="INSERT INTO Productos VALUES (NULL, '"+nombre.get()+"', "+cant.get()+")"
            queryAlumnos(query)
            mensaje['text']="El Producto "+nombre.get()+" se ha insertado exitosamente"
            nombre.delete(0,END)
            cant.delete(0,END)
            nombre.focus()
        else:
            mensaje['text']="Los campos no pueden estar vacias humano tonto"
        mostrarDatos()

    def doubleClickTabla(event):
        global claveVieja
        claveVieja=str(treeview.item(treeview.selection())["values"][0])
        nombre.delete(0,END)
        cant.delete(0,END)
        crear["state"]="disable"
        editar["state"]="normal"
        borrar["state"]="normal"
        nombre.insert(0,str(treeview.item(treeview.selection())["text"]))
        cant.insert(0,str(treeview.item(treeview.selection())["values"][0]))

    def editarRegistro():
        if len(nombre.get())!=0 and len(cant.get())!=0:
            query="UPDATE `Productos` set nombre='"+nombre.get()+"',cant="+cant.get()+" where cant="+claveVieja+"; "
            queryAlumnos(query)
            mensaje['text']="El Producto "+nombre.get()+" se a actualizado exitosamente"
            nombre.delete(0,END)
            cant.delete(0,END)
            nombre.focus()
        else:
            mensaje['text']="Los campos no pueden estar vacias humano tonto"
        mostrarDatos()
        crear["state"]="normal"
        editar["state"]="disabled"
        borrar["state"]="disabled"

    def borrarRegistro():
        query="delete from Productos where cant='"+claveVieja+"'"
        queryAlumnos(query)
        mensaje['text']="El Producto "+nombre.get()+" se a borrado exitosamente"
        nombre.delete(0,END)
        cant.delete(0,END)
        nombre.focus()
        mostrarDatos()
        crear["state"]="normal"
        editar["state"]="disabled"
        borrar["state"]="disabled"        

    treeview = ttk.Treeview(frame2, columns=2)
    treeview.grid(row=0, column=0, sticky="nsew", pady=5, padx=5)
    treeview.bind("<Double-Button-1>", doubleClickTabla)
    treeview.heading("#0", text="Nombre", anchor=CENTER)
    treeview.heading("#1", text="Cantidad", anchor=CENTER)

    vsb = ttk.Scrollbar(frame2, orient="vertical", command=treeview.yview)
    vsb.grid(row=0, column=1, sticky="nsw", pady=5, padx=5)
    hsb = ttk.Scrollbar(frame2, orient="horizontal", command=treeview.xview)
    hsb.grid(row=1, column=0, sticky="new", pady=5, padx=5)

    treeview.configure(xscrollcommand=hsb.set, yscrollcommand=vsb.set)

    crear=Button(frame3,text="Crear Producto",command=agregarRegistro)
    crear.grid(row=2,columnspan=2,sticky=W+E,pady=2, padx=2)
    crear.config(bg="#404040")
    crear.config(fg="white")
    crear.config(font=("SourceSansPro-Regular", 14), justify="center")
    editar=Button(frame3,text="Editar Producto",command=editarRegistro)
    editar.grid(row=3,columnspan=2,sticky=W+E,pady=2, padx=2)
    editar["state"]="disabled"
    editar.config(bg="#404040")
    editar.config(fg="white")
    editar.config(font=("SourceSansPro-Regular", 14), justify="center")
    borrar=Button(frame3,text="Borrar Producto",command=borrarRegistro)
    borrar.grid(row=4,columnspan=2,sticky=W+E,pady=2, padx=2)
    borrar["state"]="disabled"
    borrar.config(bg="#404040")
    borrar.config(fg="white")
    borrar.config(font=("SourceSansPro-Regular", 14), justify="center")

    def enviarlista():
        treeview_children = treeview.get_children()
        datos=[]
        t=0
        print(treeview_children)
        for item in treeview_children:
            #variable = tabla1.item("I001")["values"][0]
            text = treeview.item(item, option="text")
            value= treeview.item(item, option="values")
            print(text)
            print(value[0])
            dot=(text,value[0])
            datos.append(dot)
            listaenv1=(datos)
            print(listaenv1)
        excel_lista(listaenv1)

    listt=Button(frame3,text="Lista \n terminada",command=enviarlista)
    listt.grid(row=2,rowspan=3,column=2,sticky=W+E,pady=2, padx=45)
    listt.config(bg="#3BA25D")
    listt.config(fg="white")
    listt.config(font=("SourceSansPro-Regular", 14), justify="center")

    # print(treeview.set(item1))
    # text = treeview.item(item1, option="values")
    # print(text)


    #treeview.insert(item, tk.END, text="Subelemento 1")

#git clone https://github.com/Exusai/bodeg-on
#https://github.com/Unity-Technologies/ROS-TCP-Endpoint
#catkin_make
#roslaunch ros_tcp_endpoint endpoint.launch

    # excel_document = openpyxl.Workbook() #crear un archivo
    # excel_document = openpyxl.load_workbook('ProductosPrincipal.xlsx') #abrir documento
    # excel_document.get_sheet_names() #obtener nombres de hojas de excel
    # excel_document.get_sheet_by_name() #acceder a una hoja con su nombre

    # sheet = excel_document.get_sheet_by_name('Sheet1') #obtener el valor de una celda
    # print sheet['A2'].value

    # sheet.cell(row = 5, column = 2).value #obtener el valor de una celda

    # multiple_cells = sheet['A1':'B3']  #acceder a un rango de celdas
    # for row in multiple_cells:
    #     for cell in row:
    #         print cell.value


    # all_rows = sheet.rows #acceder a todas las filas
    # print all_rows[:]

    # all_columns = sheet.columns #acceder a todas las columnas
    # print all_columns[:]

    # sheet["A1"] = 10 #asignar valor a una celda
    # b1 = hoja.cell(row=1, column=2, value=20) #asignar valor a una celda

    # a1 = hoja["A1"] #imprimir valor de celda
    # print(a1.value)

    # c1 = hoja.cell(row=1, column=3) #actualizar valor de celda
    # c1.value = 30
    mostrarDatos()
    raiz.mainloop()